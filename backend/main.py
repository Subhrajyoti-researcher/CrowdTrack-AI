import asyncio
import io
import logging
import uuid
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import aiofiles
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from crowd_counter import process_video

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
app = FastAPI(title="CrowdTrack AI", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

OUTPUT_DIR = Path(__file__).parent / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {".wmv", ".mp4", ".avi", ".mkv", ".mov", ".m4v"}
MAX_FILE_SIZE_MB = 2000  # 2 GB soft limit

jobs: dict = {}          # job_id -> { status, progress, results?, error? }
latest_frames: dict = {} # job_id -> latest JPEG bytes for MJPEG stream
executor = ThreadPoolExecutor(max_workers=2)


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.get("/api/health")
async def health():
    return {"status": "ok"}


async def _handle_upload(file: UploadFile, mode: str) -> dict:
    """Shared upload handler — saves the file and starts a background job."""
    suffix = Path(file.filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{suffix}'. Accepted: {', '.join(ALLOWED_EXTENSIONS)}",
        )

    job_id    = str(uuid.uuid4())
    dest_path = UPLOAD_DIR / f"{job_id}{suffix}"

    async with aiofiles.open(dest_path, "wb") as out:
        chunk_size = 1024 * 1024
        while chunk := await file.read(chunk_size):
            await out.write(chunk)

    logger.info(f"[{job_id}] Saved upload → {dest_path} (mode={mode})")
    jobs[job_id] = {"status": "processing", "progress": 0, "partial_intervals": []}

    loop = asyncio.get_event_loop()
    loop.run_in_executor(executor, _run_job, job_id, str(dest_path), mode)

    return {"job_id": job_id}


@app.post("/api/upload")
async def upload_standard(file: UploadFile = File(...)):
    """Standard-mode upload — balanced precision for low-to-medium density crowds."""
    return await _handle_upload(file, mode='standard')


@app.post("/api/upload-dense")
async def upload_dense(file: UploadFile = File(...)):
    """Dense-mode upload — 99 % recall target for high-density crowds."""
    return await _handle_upload(file, mode='dense')



@app.get("/api/status/{job_id}")
async def get_status(job_id: str):
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    return jobs[job_id]


@app.get("/api/video/{job_id}")
async def get_video(job_id: str):
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    job = jobs[job_id]
    if job.get("status") != "completed":
        raise HTTPException(status_code=404, detail="Video not ready yet")
    video_path = job.get("results", {}).get("video_path")
    if not video_path or not Path(video_path).exists():
        raise HTTPException(status_code=404, detail="Annotated video not available")
    return FileResponse(
        path=video_path,
        media_type="video/mp4",
        filename=f"crowdtrack_{job_id}.mp4",
    )


@app.get("/api/stream/{job_id}")
async def stream_video(job_id: str):
    """MJPEG stream of annotated frames as they are detected."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")

    async def mjpeg_generator():
        last_sent = None
        while True:
            job = jobs.get(job_id)
            if not job:
                break

            frame_bytes = latest_frames.get(job_id)
            if frame_bytes is not None and frame_bytes is not last_sent:
                last_sent = frame_bytes
                yield (
                    b"--frame\r\n"
                    b"Content-Type: image/jpeg\r\n\r\n" +
                    frame_bytes +
                    b"\r\n"
                )

            if job.get("status") in ("completed", "error"):
                break

            await asyncio.sleep(0.15)

    return StreamingResponse(
        mjpeg_generator(),
        media_type="multipart/x-mixed-replace; boundary=frame",
    )


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1E2530")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SUMMARY_FILL = PatternFill("solid", fgColor="2A3244")
_SUMMARY_FONT = Font(bold=True, color="E2E8F0")


def _set_header_row(ws, row, values):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def _crowd_level(count, peak):
    if peak == 0:
        return "Low"
    ratio = count / peak
    if ratio >= 0.70:
        return "High"
    if ratio >= 0.35:
        return "Medium"
    return "Low"


def _build_excel(std_results, dense_results) -> bytes:
    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    labels = ["Metric", "Standard Analysis", "Dense Analysis"] if (std_results and dense_results) else \
             ["Metric", "Standard Analysis"] if std_results else ["Metric", "Dense Analysis"]
    _set_header_row(ws_sum, 1, labels)

    def _fmt_dur(secs):
        if secs is None:
            return "—"
        m, s = divmod(int(secs), 60)
        return f"{m}m {s}s"

    rows = [
        ("Video Duration",   _fmt_dur(std_results.get("duration") if std_results else None),
                             _fmt_dur(dense_results.get("duration") if dense_results else None)),
        ("FPS",              f'{std_results["fps"]:.2f}' if std_results else "—",
                             f'{dense_results["fps"]:.2f}' if dense_results else "—"),
        ("Peak Count",       std_results["overall_max"] if std_results else "—",
                             dense_results["overall_max"] if dense_results else "—"),
        ("Avg per Window",   std_results["overall_avg"] if std_results else "—",
                             dense_results["overall_avg"] if dense_results else "—"),
        ("30-s Windows",     len(std_results["intervals"]) if std_results else "—",
                             len(dense_results["intervals"]) if dense_results else "—"),
        ("Processing Time",  f'{std_results["processing_time_s"]}s' if std_results else "—",
                             f'{dense_results["processing_time_s"]}s' if dense_results else "—"),
    ]

    for r_idx, row_data in enumerate(rows, 2):
        values = row_data[:len(labels)]
        for c_idx, val in enumerate(values, 1):
            cell = ws_sum.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 1:
                cell.font = Font(bold=True)
    ws_sum.freeze_panes = "A2"
    _autosize_columns(ws_sum)

    # ── Per-mode interval sheets ────────────────────────────────────────
    def _add_interval_sheet(results, sheet_name):
        ws = wb.create_sheet(sheet_name)
        headers = ["#", "Time Window", "Min", "Avg", "Max", "Samples", "Crowd Level"]
        _set_header_row(ws, 1, headers)

        peak = results["overall_max"]
        intervals = results["intervals"]

        # Summary row
        overall_min = min(i["min_count"] for i in intervals) if intervals else 0
        total_samples = sum(i["samples"] for i in intervals)
        summary_vals = ["Overall", intervals[0]["label"].split("–")[0].strip() + " – " +
                        intervals[-1]["label"].split("–")[-1].strip() if len(intervals) > 1 else (intervals[0]["label"] if intervals else "—"),
                        overall_min, results["overall_avg"], peak, total_samples,
                        _crowd_level(peak, peak) + " (Peak)"]
        for c_idx, val in enumerate(summary_vals, 1):
            cell = ws.cell(row=2, column=c_idx, value=val)
            cell.fill = _SUMMARY_FILL
            cell.font = _SUMMARY_FONT

        # Per-window rows
        for r_idx, interval in enumerate(intervals, 3):
            ws.cell(row=r_idx, column=1, value=r_idx - 2)
            ws.cell(row=r_idx, column=2, value=interval["label"])
            ws.cell(row=r_idx, column=3, value=interval["min_count"])
            ws.cell(row=r_idx, column=4, value=interval["avg_count"])
            ws.cell(row=r_idx, column=5, value=interval["max_count"])
            ws.cell(row=r_idx, column=6, value=interval["samples"])
            ws.cell(row=r_idx, column=7, value=_crowd_level(interval["max_count"], peak))

        ws.freeze_panes = "A3"
        _autosize_columns(ws)

    if std_results:
        _add_interval_sheet(std_results, "Standard Analysis")
    if dense_results:
        _add_interval_sheet(dense_results, "Dense Analysis")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.get("/api/export-excel")
async def export_excel(
    std_job_id: str | None = Query(default=None),
    dense_job_id: str | None = Query(default=None),
):
    std_results = None
    dense_results = None

    if std_job_id:
        job = jobs.get(std_job_id)
        if not job or job.get("status") != "completed":
            raise HTTPException(status_code=404, detail="Standard job not completed")
        std_results = job["results"]

    if dense_job_id:
        job = jobs.get(dense_job_id)
        if not job or job.get("status") != "completed":
            raise HTTPException(status_code=404, detail="Dense job not completed")
        dense_results = job["results"]

    if not std_results and not dense_results:
        raise HTTPException(status_code=400, detail="No job IDs provided")

    xlsx_bytes = _build_excel(std_results, dense_results)

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=crowdtrack_results.xlsx"},
    )


# ---------------------------------------------------------------------------
# Background worker
# ---------------------------------------------------------------------------

def _run_job(job_id: str, file_path: str, mode: str = 'standard'):
    import time
    output_path = str(OUTPUT_DIR / f"{job_id}_annotated.mp4")
    try:
        logger.info(f"[{job_id}] Processing started (mode={mode})")
        t_start = time.time()

        def on_progress(pct: int):
            jobs[job_id]["progress"] = pct

        def on_frame(frame_bytes: bytes):
            latest_frames[job_id] = frame_bytes

        def on_window(interval: dict):
            jobs[job_id]["partial_intervals"].append(interval)

        # Dense mode uses ~4x more tiles per frame; sample every 2 s
        # to keep wall-clock time comparable to standard mode.
        sample_secs = 2.0 if mode == 'dense' else 1.0

        result = process_video(
            file_path,
            progress_callback=on_progress,
            output_video_path=output_path,
            frame_callback=on_frame,
            window_callback=on_window,
            sample_every_n_seconds=sample_secs,
            mode=mode,
        )

        result["processing_time_s"] = round(time.time() - t_start, 1)

        # Attach video URL if output was written successfully
        if result.get("video_path") and Path(result["video_path"]).exists():
            result["video_url"] = f"/api/video/{job_id}"

        jobs[job_id] = {"status": "completed", "progress": 100, "results": result}
        logger.info(f"[{job_id}] Completed — {len(result['intervals'])} windows | {result['processing_time_s']}s")

    except Exception as exc:
        logger.exception(f"[{job_id}] Failed")
        error_msg = str(exc) or f"{type(exc).__name__}: processing failed"
        jobs[job_id] = {"status": "error", "progress": 0, "error": error_msg}
    finally:
        # Remove uploaded file after processing
        try:
            Path(file_path).unlink(missing_ok=True)
        except Exception:
            pass
        # Clear frame buffer
        latest_frames.pop(job_id, None)


# ---------------------------------------------------------------------------
# Serve frontend (must be last — catch-all)
# ---------------------------------------------------------------------------

FRONTEND_DIR = Path(__file__).parent.parent / "react-frontend" / "dist"
if FRONTEND_DIR.exists():
    app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="static")
